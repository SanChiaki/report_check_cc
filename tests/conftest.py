import pytest
from pathlib import Path
import openpyxl
from PIL import Image
from io import BytesIO


@pytest.fixture
def sample_excel_path(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "报告"
    ws["A1"] = "交付报告"
    ws["A2"] = "项目名称"
    ws["B2"] = "XXX服务器部署项目"
    ws["A3"] = "报告日期"
    ws["B3"] = "2026-03-22"
    ws["A5"] = "交付内容"
    ws["A6"] = "1. 服务器部署"
    ws["A7"] = "2. 网络配置"
    ws["A10"] = "移交记录"
    ws["A11"] = "移交人"
    ws["B11"] = "张三"
    ws["A12"] = "移交时间"
    ws["B12"] = "2026-03-20"
    ws["A13"] = "移交命令"
    ws["B13"] = "deploy --all"

    img = Image.new("RGB", (100, 100), color="white")
    buf = BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    xl_img = openpyxl.drawing.image.Image(buf)
    xl_img.anchor = "A15"
    ws.add_image(xl_img)

    path = tmp_path / "test_report.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def sample_excel_no_images(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "简单报告"
    ws["A1"] = "交付内容"
    ws["A2"] = "测试数据"
    path = tmp_path / "simple_report.xlsx"
    wb.save(path)
    return path
