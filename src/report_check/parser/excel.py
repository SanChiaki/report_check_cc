import logging
from pathlib import Path
from typing import TYPE_CHECKING
import openpyxl
from openpyxl.utils import get_column_letter
from report_check.parser.base import BaseParser
from report_check.parser.models import ContentBlock, ImageData, ReportData
from report_check.parser.utils import detect_and_convert_format

if TYPE_CHECKING:
    from report_check.storage.artifacts import TaskArtifacts

logger = logging.getLogger(__name__)
NEARBY_RADIUS = 3


class ExcelParser(BaseParser):
    def __init__(self, artifacts: "TaskArtifacts | None" = None):
        """
        Args:
            artifacts: Optional TaskArtifacts instance for recording parse details
        """
        self.artifacts = artifacts

    def parse(self, file_path: str) -> ReportData:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        content_blocks = self._extract_cells(ws)
        images = self._extract_images(ws, content_blocks)

        parse_metadata = {
            "type": "excel",
            "sheet_name": ws.title,
            "row_count": ws.max_row or 0,
            "col_count": ws.max_column or 0,
            "content_block_count": len(content_blocks),
            "image_count": len(images),
        }

        if self.artifacts:
            self.artifacts.save_parse_metadata(parse_metadata)

        return ReportData(
            file_name=Path(file_path).name,
            source_type="excel",
            content_blocks=content_blocks,
            images=images,
            metadata={"sheet_name": ws.title, "row_count": ws.max_row or 0, "col_count": ws.max_column or 0},
        )

    def _extract_cells(self, ws) -> list[ContentBlock]:
        blocks = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    data_type = cell.data_type or "s"
                    content_type = "number" if data_type == "n" else "date" if data_type == "d" else "text"
                    blocks.append(ContentBlock(
                        content=str(cell.value),
                        location=cell.coordinate,
                        content_type=content_type,
                        metadata={"row": cell.row, "col": cell.column, "excel_type": data_type},
                    ))
        return blocks

    def _extract_images(self, ws, content_blocks: list[ContentBlock]) -> list[ImageData]:
        images = []
        for i, img in enumerate(ws._images):
            try:
                image_data = self._get_image_bytes(img)
                if image_data is None:
                    continue
                fmt = detect_and_convert_format(image_data)
                if fmt is None:
                    continue
                anchor = self._get_anchor(img)
                nearby = self._get_nearby_blocks(content_blocks, anchor.get("row", 1), anchor.get("col", 1))

                final_data = image_data if fmt[1] is None else fmt[1]
                final_format = fmt[0]

                image_id = f"img_{i}"
                images.append(ImageData(
                    id=image_id,
                    data=final_data,
                    format=final_format,
                    anchor=anchor,
                    nearby_blocks=nearby,
                ))

                # Save image to artifacts
                if self.artifacts:
                    self.artifacts.save_parsed_image(
                        image_id=image_id,
                        data=final_data,
                        format=final_format,
                        metadata={
                            "anchor": anchor,
                            "nearby_blocks_count": len(nearby),
                        }
                    )
            except Exception as e:
                logger.warning(f"Failed to extract image {i}: {e}")
        return images

    def _get_image_bytes(self, img) -> bytes | None:
        try:
            if hasattr(img, "_data"):
                return img._data()
            if hasattr(img, "ref"):
                with open(img.ref, "rb") as f:
                    return f.read()
        except Exception as e:
            logger.warning(f"Cannot read image data: {e}")
        return None

    def _get_anchor(self, img) -> dict:
        anchor = {}
        try:
            if hasattr(img.anchor, "_from"):
                anchor_cell = img.anchor._from
                row = anchor_cell.row + 1
                col = anchor_cell.col + 1
                anchor = {"row": row, "col": col, "cell_ref": f"{get_column_letter(col)}{row}"}
            elif isinstance(img.anchor, str):
                from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
                col_letter, row = coordinate_from_string(img.anchor)
                col = column_index_from_string(col_letter)
                anchor = {"row": row, "col": col, "cell_ref": img.anchor}
        except Exception as e:
            logger.warning(f"Cannot parse anchor: {e}")
            anchor = {"row": 1, "col": 1, "cell_ref": "A1"}
        return anchor

    def _get_nearby_blocks(self, content_blocks: list[ContentBlock], row: int, col: int) -> list[ContentBlock]:
        nearby = []
        for block in content_blocks:
            block_row = block.metadata.get("row", 0)
            block_col = block.metadata.get("col", 0)
            if (abs(block_row - row) <= NEARBY_RADIUS and
                abs(block_col - col) <= NEARBY_RADIUS):
                nearby.append(block)
        return nearby
