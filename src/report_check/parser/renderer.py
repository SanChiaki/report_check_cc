"""Report renderer for converting reports to images."""

import logging
from typing import TYPE_CHECKING, List
from io import BytesIO

if TYPE_CHECKING:
    from report_check.parser.report_data import ReportData
    from report_check.storage.artifacts import CheckArtifact

logger = logging.getLogger(__name__)


class ReportRenderer:
    """Render reports as images for multimodal analysis."""

    async def render(self, report_data: "ReportData", artifacts: "CheckArtifact | None" = None) -> List[bytes]:
        """Render report to images.

        Args:
            report_data: Report data to render
            artifacts: Optional artifacts manager for saving rendered images

        Returns:
            List of image bytes (one per page)
        """
        if report_data.source_type == "pdf":
            return await self._render_pdf(report_data, artifacts)
        elif report_data.source_type == "excel":
            return await self._render_excel(report_data, artifacts)
        else:
            logger.error(f"Unsupported source type: {report_data.source_type}")
            return []

    async def _render_pdf(self, report_data: "ReportData", artifacts: "CheckArtifact | None") -> List[bytes]:
        """Render PDF pages as images."""
        images = []

        # Check if we already have page images from parsing
        if hasattr(report_data, 'page_images') and report_data.page_images:
            for page_num, img_data in report_data.page_images.items():
                images.append(img_data)
                if artifacts:
                    artifacts.add_image_evidence(f"page_{page_num}", img_data, "png")
            return images

        # Otherwise render from PDF file
        import fitz  # PyMuPDF

        pdf_path = report_data.metadata.get("file_path")
        if not pdf_path:
            return []

        doc = fitz.open(pdf_path)
        for page_num in range(len(doc)):
            page = doc[page_num]
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))  # 2x scale
            img_data = pix.tobytes("png")
            images.append(img_data)

            if artifacts:
                artifacts.add_image_evidence(f"page_{page_num+1}", img_data, "png")

        doc.close()
        return images

    async def _render_excel(self, report_data: "ReportData", artifacts: "CheckArtifact | None") -> List[bytes]:
        """Render Excel as image."""
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as XLImage
        from PIL import Image, ImageDraw, ImageFont
        import io
        from pathlib import Path

        excel_path = report_data.metadata.get("file_path")
        if not excel_path:
            return []

        # Check if file exists
        if not Path(excel_path).exists():
            logger.warning(f"Excel file not found: {excel_path}")
            return []

        try:
            wb = load_workbook(excel_path)
        except Exception as e:
            logger.error(f"Failed to load Excel file: {e}")
            return []

        images = []

        for sheet in wb.worksheets:
            # Create blank canvas
            img = Image.new('RGB', (2000, 3000), 'white')
            draw = ImageDraw.Draw(img)

            # Draw cell content
            y_offset = 10
            for row in sheet.iter_rows(values_only=True):
                text = " | ".join(str(cell) if cell else "" for cell in row)
                if text.strip():
                    draw.text((10, y_offset), text[:100], fill='black')
                    y_offset += 20

            # Crop to content
            bbox = img.getbbox()
            if bbox:
                img = img.crop(bbox)

            # Convert to bytes
            buf = io.BytesIO()
            img.save(buf, format='PNG')
            img_data = buf.getvalue()
            images.append(img_data)

            if artifacts:
                artifacts.add_image_evidence(f"sheet_{sheet.title}", img_data, "png")

        return images
